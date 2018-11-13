import base64
import json
import logging
import sys
import traceback

import boto3
import botocore

logging.basicConfig()
log = logging.getLogger()
log.setLevel(logging.INFO)

sys.path.insert(0, './pip')
sys.path.insert(0, './pip_openpyxl')

try:
  print(xlsxwriter)
except NameError:
  import xlsxwriter
except:
  print("Unexpected error:", sys.exc_info()[0])

try:
  print(openpyxl)
  # problem: lose images and charts
except NameError:
  from openpyxl import load_workbook
except:
  print("Unexpected error:", sys.exc_info()[0])

s3 = boto3.resource("s3")


def get_file_from_s3(bucket_name, file_name):

  # KEY = 'Y2TSL8HJY_ManageMashupTokenIn.wsdl' # replace with your object key

  client_s3 = boto3.client("s3")

  s3 = boto3.resource('s3')

  local_wsdl_location = '/tmp/{}'.format(file_name)

  try:
    # log.info(client_s3.list_objects_v2(Bucket=bucket_name))
    s3.meta.client.download_file(bucket_name, file_name, local_wsdl_location)
    # client_s3.get_object(Bucket=bucket_name, Key=KEY)#, 'wsdl/{}'.format(KEY))
  except botocore.exceptions.ClientError as e:
    # if e.response['Error']['Code'] == "404":
    str_msg = traceback.format_exc().splitlines()
    log.error("Error: During downloading file - {0}".format(str_msg))
    return {"error": str_msg}

  return local_wsdl_location


def handler(event, context):
  print(event)
  q = getQuery(event)
  print(q)

  filename = q.get("filename", None)
  if filename is None:
    return "filename is required"

  saveToPath = "/tmp/"
  saved_location = "{}converted-{}".format(saveToPath, filename)

  file_location = get_file_from_s3("sheet-bsg.support", filename)
  print(file_location)
  if "error" in file_location:
    return file_location

  wb = load_workbook(file_location)
  ws = wb.active
  ws['A6'] = "fjfj"
  wb.save(saved_location)

  s3_converted_location = upload_to_s3(
      "sheet-bsg.support", saved_location, "converted/{}".format(filename))

  return {"url": s3_converted_location}

  # excel_file = ""
  # with open(saved_location,'rb') as f:
  #   excel_file = f.read()
  #   excel_file = base64.b64encode(excel_file).decode("utf-8")
  #
  # return {
  #     "statusCode": 200,
  #     'isBase64Encoded': True,
  #     'headers': {
  #         'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
  #         'Access-Control-Allow-Origin': '*',
  #         'Content-Disposition': 'attachment; filename="test.xlsx"'
  #     },
  #     "body": json.dumps(excel_file)#["queryStringParameters"]
  #     }


def handler_xlsxwriter(event, context):
  print(event)
  q = getQuery(event)

  return_headers = {"Access-Control-Allow-Origin": "*",
                    "Access-Control-Expose-Headers": ""}

  files_suffix = q.get("files_suffix", "")
  file_name = q.get("file_name", None)
  columns = q.get("columns", None)
  arrdata = q.get("arrdata", None)
  if arrdata:
    if file_name is None:
      return {
          'statusCode': 500,
          'headers': return_headers,
          'body': json.dumps({"errorMessage": "file_name is required"}),
      }
    if columns is None:
      return {
          'statusCode': 500,
          'headers': return_headers,
          'body': json.dumps({"errorMessage": "columns is required"}),
      }
  else:
    return {
        'statusCode': 500,
        'headers': return_headers,
        'body': json.dumps({"errorMessage": "arrdata is required"}),
    }

  saveToPath = "/tmp/"
  name_with_suffix = "{0}_{1}".format(file_name, files_suffix)
  workbook = xlsxwriter.Workbook(
      saveToPath + '{}.xlsx'.format(name_with_suffix))
  worksheet = workbook.add_worksheet()

  last_row = len(arrdata)
  last_col = len(columns) - 1
  worksheet.add_table(0, 0, last_row, last_col, {
                      'data': arrdata, 'columns': columns})
  # worksheet.add_table(0,0,1,2,{'data': [[1,2]],'columns':[{"header":"Column 1"},{"header":"Column 2"},{"header":"Column 3"}]})

  workbook.close()

  result = {"url": upload_to_s3(
      "sheet-bsg.support", name_with_suffix, "jenax/pcr/files/")}

  return {
      'statusCode': 200,
      'headers': return_headers,
      'body': json.dumps(result),
  }

  # excel_file = ""
  # with open(saveToPath+'hello.xlsx','rb') as f:
  #   excel_file = f.read()
  #   excel_file = base64.b64encode(excel_file).decode("utf-8")
  #
  # return {
  #     "statusCode": 200,
  #     'isBase64Encoded': True,
  #     'headers': {
  #         'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
  #         'Access-Control-Allow-Origin': '*',
  #         'Content-Disposition': 'attachment; filename="hello.xlsx"'
  #     },
  #     "body": json.dumps(excel_file)#["queryStringParameters"]
  #     }


def upload_to_s3(bucket_name, local_filename, key_name):

  # with open('/tmp/{}'.format(temp_filename), 'rb') as csvfile:
  #  res = s3.Bucket(bucket_name).put_object(ACL="public-read", Key="{}".format(key_name), Body=csvfile)
  s3.meta.client.upload_file(
      local_filename, bucket_name, key_name)

  # bucket_location = boto3.client('s3').get_bucket_location(Bucket=bucket_name)
  # object_url = "https://s3-{0}.amazonaws.com/{1}/{2}".format(
  #  bucket_location['LocationConstraint'],
  #  bucket_name,
  #  key_name)
  object_url = "https://downloadsheet.bsg.support/{0}".format(key_name)

  return object_url


def getQuery(e):
  try:
    method = e["httpMethod"]
  except:
    method = e["context"]["http-method"]

  if method == "GET":
    try:
      return e["queryStringParameters"]
    except:
      return e["params"]["querystring"]
  else:
    if "body" in e:
      return json.loads(e["body"])
    else:
      return e["body-json"]
